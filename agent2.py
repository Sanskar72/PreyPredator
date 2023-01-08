# Imports
from genenvironment import genEnvironment, spawnCreatures, preyMovement, predatorMovement, BFS
import pandas as pd
from openpyxl import load_workbook
import random
import copy

def simulateFuture(nodes, agentPos, preyPos, agentPreyDist):
    """_summary_
        The Function is used to simulate the entire process after surveying for belief states of each nodes. we simulate to find out the next possible 
        cases of our system and out of these 10 options we choose the best option for our agent to consider
    Args:
        nodes (Dictionary): Dictionary with all the node information in the graph
        predatorPos (int): The location of the Predator on the chain graph
        agentPos (int): The location of the Agent on the chain graph
        preyPos (int): The location of the Prey on the chain graph

    Returns: It returns the Agent position after simulating the entire process 10 times and then picking one at random
        _type_: int
    """
    nextNeigh = list()
    simulationCount = 10
    for _ in range(simulationCount):
        simulatedAgentPos = copy.deepcopy(agentPos)
        simulatedPreyPos = copy.deepcopy(preyPos)
        timeStamp = 0
        while timeStamp<agentPreyDist and simulatedAgentPos != simulatedPreyPos:
            maybeNeigh = [simulatedAgentPos]
            # Find the path between the Neighbour and the Prey
            agentPathPrey = random.choice(BFS(nodes, simulatedAgentPos, simulatedPreyPos)["path"])
            agentPreyDist = len(agentPathPrey)
            maxagentPreyNeighDiff = 0
            for neighbour in nodes[simulatedAgentPos]["neighbours"]:
                preyDict = BFS(nodes, neighbour, simulatedPreyPos)
                randInd = random.randint(0,len(preyDict["path"])-1)
                newPathPrey = preyDict["path"][randInd]
                agentPreyNeighDiff0 = (agentPreyDist-len(newPathPrey))
                # If agentPreyNeighDiff0 distance is greater than before clear and update
                if agentPreyNeighDiff0>maxagentPreyNeighDiff:
                    maybeNeigh.clear()
                    maxagentPreyNeighDiff = agentPreyNeighDiff0
                    maybeNeigh.append(neighbour)
                # Node to keep the distance between Predator and Agent same
                if agentPreyNeighDiff0==maxagentPreyNeighDiff:
                    maybeNeigh.append(neighbour)
            simulatedAgentPos = random.choice(maybeNeigh)
            simulatedPreyPos = preyMovement(nodes, simulatedPreyPos)
            timeStamp += 1
        nextNeigh.append(simulatedAgentPos)
    finalAgentPos = random.choice(nextNeigh)
    agentPath = random.choice(BFS(nodes, agentPos, finalAgentPos)["path"])
    
    if len(agentPath) > 1:
        agentPos = agentPath[1]
    return agentPos

def agent2Movement(nodes, predPos, agentPos, preyPos, probUse, distUse):
    """_summary_
        This Function is used to plan the Agent 2 movement taking into account the simulations as well as the distance 
        between the prey with the Agent and the predator with the Agent
    Args:
        nodes (Dictionary): Dictionary with all the node information in the graph
        predPos (_type_): _description_
        predatorPos (int): The location of the Predator on the chain graph
        agentPos (int): The location of the Agent on the chain graph
        preyPos (int): The location of the Prey on the chain graph
        distUse (int): Number of times the agent uses distance i.e moves away from the predator when agentPreyDist > agentPredDist

    Returns: The final agent Position for the next move, the number of times probability is Used and simulations take place and if distance 
    is used to make the agent move away from the predator
        _type_: int
    """
    maxPredDistNeigh = [agentPos]
    agentPathPred = random.choice(BFS(nodes, agentPos, predPos)["path"])
    agentPathPrey = random.choice(BFS(nodes, agentPos, preyPos)["path"])
    agentPredDist = len(agentPathPred)
    agentPreyDist = len(agentPathPrey)
    maxPredDist = agentPredDist
    if agentPreyDist < agentPredDist:
        probUse += 1
        agentPos = simulateFuture(nodes, agentPos, preyPos, agentPreyDist)
    else:
        
        distUse += 1
        for neighbour in nodes[agentPos]["neighbours"]:
            #print(neighbour, " maxPredDistNeigh", maxPredDistNeigh)
            predatorDict = BFS(nodes, neighbour, predPos)
            # Find the distance between the Neighbour and the Predator
            randInd = random.randint(0,len(predatorDict["path"])-1)
            newPathPred = predatorDict["path"][randInd]
            
            if len(newPathPred) > maxPredDist:
                maxPredDistNeigh.clear()
                maxPredDist = len(newPathPred)
                maxPredDistNeigh.append(neighbour)
                
            elif len(newPathPred) == maxPredDist:
                maxPredDistNeigh.append(neighbour)
        # Agent moves away from predator near to the prey
        agentPos = random.choice(maxPredDistNeigh)
    return agentPos, probUse, distUse

def agent2(nodes, predatorPos, agentPos, preyPos):
    """_summary_
        Main function for the running of Agent 2 prioritizing to maximize the distance between the agent and the predator and minimize 
        the distance between the Agent and Prey
    Args:
        nodes (Dictionary): Dictionary with all the node information in the graph
        predatorPos (int): The location of the Predator on the chain graph
        agentPos (int): The location of the Agent on the chain graph
        preyPos (int): The location of the Prey on the chain graph

    Returns: The status of completion of the agent the final step count for the Agent to reach the goal, the path taken by the agent 2 
            and the path of the predator.
        _type_: json
    """
    threshold = 1000
    agentPath = list()
    agentPath.append(agentPos)
    predPath = list()
    predPath.append(predatorPos)
    probUse, distUse = 0, 0
    for counter in range(1,threshold+1):
        # print("Counter: ", counter)
        # print("Agent: ", agentPos)
        # print("Prey: ", preyPos)
        # print("Predator: ", predatorPos)
        # print("========================")
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "probUse":probUse, "distUse":distUse}
        
        if agentPos == predatorPos:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "probUse":probUse, "distUse":distUse}
        #print(agentPos)
        agentPos, probUse, distUse = agent2Movement(nodes, predatorPos, agentPos, preyPos, probUse, distUse)
        #print("ss=", agentPos)
        #ELSE STAY SAME PLACE.
        agentPath.append(agentPos)
        # If Agent reaches Prey Position which ih the Goal State
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "probUse":probUse, "distUse":distUse}
        # Making the prey move
        preyPos = preyMovement(nodes, preyPos)
        # After the Prey movement takes place
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "probUse":probUse, "distUse":distUse}
        # Conditions for Predator killing the agent
        predDict = predatorMovement(agentPos, predatorPos, nodes)
        if predDict["statusCode"] == 200:
            predatorPos = predDict["predatorPos"]
            predPath.append(predatorPos)
            
        elif predDict["statusCode"] == 400:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "probUse":probUse, "distUse":distUse}
        
    return {"statusCode": 404, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "probUse":probUse, "distUse":distUse}
        
def driver():
    """_summary_
        Driver Code for the Agent 2
    """
    nodes, _ = genEnvironment()
    predatorPos, agentPos, preyPos = spawnCreatures()
    return agent2(nodes, predatorPos, agentPos, preyPos)

def dataCollection():
    """_summary_
        Function to collect the data regarding Agent 2, its performance and all other statistical information
    """
    final_data = list()
    for i in range(1000):
        print("Counter: ",i)
        data = driver()
        final_data.append(data)
            
    df1 = pd.DataFrame(final_data)
    book = load_workbook('Agent2.xlsx')
    writer = pd.ExcelWriter('Agent2.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header = False)

    writer.save()
        
dataCollection()

def testDriver():
    # nodes, size = genEnvironment()
    # predatorPos, agentPos, preyPos = spawnCreatures()
    # #agent1data = agent1(nodes, size, predatorPos, agentPos, preyPos)     
    # #print(agent1data)
    # for i in range(10):
    #     predatorPos = predatorMovement(agentPos, predatorPos, nodes)["predatorPos"]
    data = driver()
    print("data: ",data)
#testDriver()

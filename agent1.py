# Imports
from genenvironment import genEnvironment, spawnCreatures, preyMovement, predatorMovement, BFS
import pandas as pd
from openpyxl import load_workbook
import random

def agent1(nodes, predatorPos, agentPos, preyPos):
    """_summary_
        Main function for the running of Agent 1 prioritizing to maximize the distance between the agent and the predator and minimize 
        the distance between the Agent and Prey
    Args:
        nodes (Dictionary): Dictionary with all the node information in the graph
        predatorPos (int): The location of the Predator on the chain graph
        agentPos (int): The location of the Agent on the chain graph
        preyPos (int): The location of the Prey on the chain graph

    Returns: The status of completion of the agent the final step count for the Agent to reach the goal, the path taken by the agent1 
            and the path of the predator.
        _type_: json
    """
    threshold = 1000
    agentPath = list()
    agentPath.append(agentPos)
    predPath = list()
    predPath.append(predatorPos)
    for counter in range(1,threshold+1):
        # print("Counter: ", counter)
        # print("Agent: ", agentPos)
        # print("Prey: ", preyPos)
        # print("Predator: ", predatorPos)
        # print("========================")
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath}
        
        if agentPos == predatorPos:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath}
        
        maxagentPredNeighDiff = 0
        maxagentPreyNeighDiff = 0
        nextNeigh = list()
        maxDiff = -50
        maybeNeigh = list()
        maxPredDistNeigh = list()
        agentPathPred = random.choice(BFS(nodes, agentPos, predatorPos)["path"])
        agentPathPrey = random.choice(BFS(nodes, agentPos, preyPos)["path"])
        agentPredDist = len(agentPathPred)
        agentPreyDist = len(agentPathPrey)
        maxPredDist = agentPredDist
        for neighbour in nodes[agentPos]["neighbours"]:
            # Find the distance between the Neighbour and the Prey
            preyDict = BFS(nodes, neighbour, preyPos)
            # Find the distance between the Neighbour and the Predator
            predatorDict = BFS(nodes, neighbour, predatorPos)
            randInd = random.randint(0,len(preyDict["path"])-1)
            newPathPrey = preyDict["path"][randInd]
            randInd = random.randint(0,len(predatorDict["path"])-1)
            newPathPred = predatorDict["path"][randInd]
            difference0 = len(newPathPred) - len(newPathPrey)
            agentPredNeighDiff0 = (len(newPathPred)-agentPredDist)
            agentPreyNeighDiff0 = (agentPreyDist-len(newPathPrey))
            # Get the closest neighbour to the prey for the agent to move to
            if agentPreyNeighDiff0>maxagentPreyNeighDiff and agentPredNeighDiff0>maxagentPredNeighDiff:
                maybeNeigh.clear()
                maxagentPredNeighDiff = agentPredNeighDiff0
                maxagentPreyNeighDiff = agentPreyNeighDiff0
                maybeNeigh.append(neighbour)
            # If neighbour node keeps distance to the prey and predator same as before
            if agentPreyNeighDiff0==maxagentPreyNeighDiff and agentPredNeighDiff0==maxagentPredNeighDiff:
                maybeNeigh.append(neighbour)
            # Node to increase the distance between Predator and Agent
            if len(newPathPred) > maxPredDist:
                maxPredDistNeigh.clear()
                maxPredDist = len(newPathPred)
                maxPredDistNeigh.append(neighbour)
            # Node to keep the distance between Predator and Agent same
            if len(newPathPred) == maxPredDist:
                maxPredDistNeigh.append(neighbour)
            
            if difference0>maxDiff:
                nextNeigh.clear()
                maxDiff = difference0
                nextNeigh.append(neighbour)
                
            if difference0==maxDiff:
                nextNeigh.append(neighbour)
        # Agent moves away from predator near to the prey
        if maxDiff >= 0:
            #print("maxd calld")
            agentPos = random.choice(nextNeigh)
        elif len(maybeNeigh) > 0:
            #print("maybe calld")
            agentPos = random.choice(maybeNeigh)
        elif len(maxPredDistNeigh) > 0:
            #print("max pred calld")
            agentPos = random.choice(maxPredDistNeigh)
        #ELSE STAY SAME PLACE.
        agentPath.append(agentPos)
        # If Agent reaches Prey Position which ih the Goal State
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath}
        # Making the prey move
        preyPos = preyMovement(nodes, preyPos)
        # After the Prey movement takes place
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath}
        # Conditions for Predator killing the agent
        predDict = predatorMovement(agentPos, predatorPos, nodes)
        if predDict["statusCode"] == 200:
            predatorPos = predDict["predatorPos"]
            predPath.append(predatorPos)
            
        elif predDict["statusCode"] == 400:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath}
        
    return {"statusCode": 404, "steps":counter, "AgentPath":agentPath, "PredPath":predPath}
        
def driver():
    """_summary_
        Driver Code for the Agent 1
    """
    nodes, _ = genEnvironment()
    predatorPos, agentPos, preyPos = spawnCreatures()
    return agent1(nodes, predatorPos, agentPos, preyPos)

def dataCollection():
    """_summary_
        Function to collect the data regarding Agent 1, its performance and all other statistical information
    """
    final_data = list()
    for i in range(300):
        print("Counter: ",i)
        data = driver()
        final_data.append(data)
            
    df1 = pd.DataFrame(final_data)
    book = load_workbook('Agent1.xlsx')
    writer = pd.ExcelWriter('Agent1.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header = False)

    writer.save()
        
dataCollection()

def testDriver():
    # nodes, size = genEnvironment()
    # predatorPos, agentPos, preyPos = spawnCreatures()
    # #agent1data = agent1(nodes, size, predatorPos, agentPos, preyPos)     
    # #print(agent1data)
    # for i in range(10):
    #     predatorPos = predatorMovement(agentPos, predatorPos, nodes)["predatorPos"]
    for _ in range(10):
        data = driver()
        print("dtaa", data)

#testDriver()

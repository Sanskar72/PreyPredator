# Imports
from genenvironment import genEnvironment, spawnCreatures, preyMovement, predatorMovement, BFS
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import random
import copy

def generatePreyProb(size, agentPos):
    """_summary_
        Function to initialize the Prey node position where the node with the agent has probability as 0 and the initial probability of 1/(size of graph -1) to every node 
        in the graph

    Args:
        size (int): Length of the graph 
        agentPos (int): Location of the agent on the graph

    Returns: A list of the initialized probabilities for the entire graph
        _type_: list
    """
    preyNodeProb = list()
    for i in range(size):
        if i == agentPos:
            preyNodeProb.append(0)
            continue
        # Probability Initialization
        preyNodeProb.append(1/(size-1))
    return preyNodeProb

def updateTransitPreyProb(nodes, size, preyNodeProb):
    """_summary_
        This Function is triggered after the prey moves which is done to update the belief states. Now, every node and its neighbour get updated based on the new information. The Transition probability of the node is updated taking into consideration 
        the previous initialized Probabilities and the neighbours of the node into consideration
        
    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        preyNodeProb (list): The list of the initialized probabilities for the entire graph for the prey
    
    Returns: The updated probability matrix for the entire graph maximizing the probability of the location of the prey
        _type_: list
    """
    newPreyNodeProb = [-100]*size
    for node in range(size):
        prob = preyNodeProb[node]/(nodes[node]["degree"]+1)
        for i in range(nodes[node]["degree"]):
            neighbour = nodes[node]["neighbours"][i]
            # Probability Formula used based on Conditional Probability
            prob += preyNodeProb[neighbour]/(nodes[neighbour]["degree"]+1)
            newPreyNodeProb[node] = prob
            
    return newPreyNodeProb

def updateSurveyPreyProd(size, surveySpot, preyNodeProb, preyPos):
    """_summary_
        This function in used to survey each and every node and create a belief state. After surveying the graph for the prey we keep updating the probabilities of each node based on the conditional probability
    Args:
        size (int): Length of the graph
        surveySpot (int): Location with the highest probability of having the prey used for further calculation of Agent movement
        preyNodeProb (list): The list of the initialized probabilities for the entire graph
        preyPos (int): Current location of the prey on the graph

    Returns: 
        Surveying all the nodes in the graph we geth the final probability matrix using conditional Probability to predict the prey location
     _type_: list

    """
    # Success Condition for survey
    if surveySpot == preyPos:
        for i in range(size):
            if i == surveySpot:
                preyNodeProb[i] = 1
                continue
            preyNodeProb[i] = 0
        return preyNodeProb
    else:
        newPreyNodeProb = [-100]*size
        for i in range(size):
            if i == surveySpot:
                newPreyNodeProb[i] = 0
                continue
            # Update the predator Survey probability using the previous survey and the new moves 
            newPreyNodeProb[i] = preyNodeProb[i]/(1-preyNodeProb[surveySpot])
        return newPreyNodeProb
    
def simulateFuture(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist):
    """_summary_
        The Function is used to simulate the entire process after surveying for belief states of each nodes. we simulate 
    Args:
        nodes (Dictionary): Dictionary with all the node information in the graph
        predatorPos (int): The location of the Predator on the chain graph
        agentPos (int): The location of the Agent on the chain graph
        preyPos (int): The location of the Prey on the chain graph

    Returns: It returns the Agent position after simulating the entire process 10 times and then picking one at random
        _type_: int
    """
    nextNeigh = list()
    simulationCount = 10
    for _ in range(simulationCount):
        simulatedAgentPos = copy.deepcopy(agentPos)
        simulatedPreyNodeProb = copy.deepcopy(preyNodeProb)
        simulatedPreyPos = copy.deepcopy(preyPos)
        timeStamp = 0
        while timeStamp<agentPreyDist and simulatedAgentPos != simulatedPreyPos:
            maybeNeigh = [simulatedAgentPos]
            maybePrey = np.max(simulatedPreyNodeProb)
            maybeprey = [i for i,prob in enumerate(simulatedPreyNodeProb) if prob == maybePrey]
            maybePrey = random.choice(maybeprey)
            # Survey and create the belief States for the Future Simulations
            simulatedPreyNodeProb = updateSurveyPreyProd(size, maybePrey, simulatedPreyNodeProb, simulatedPreyPos)
            maybePrey = np.max(simulatedPreyNodeProb)
            maybeprey = [i for i,prob in enumerate(simulatedPreyNodeProb) if prob == maybePrey]
            maybePrey = random.choice(maybeprey)
            # Find the path between the Neighbour and the Prey
            agentPathPrey = random.choice(BFS(nodes, simulatedAgentPos, maybePrey)["path"])
            agentPreyDist = len(agentPathPrey)
            maxagentPreyNeighDiff = 0
            for neighbour in nodes[simulatedAgentPos]["neighbours"]:
                preyDict = BFS(nodes, neighbour, maybePrey)
                randInd = random.randint(0,len(preyDict["path"])-1)
                newPathPrey = preyDict["path"][randInd]
                agentPreyNeighDiff0 = (agentPreyDist-len(newPathPrey))
                # If agentPreyNeighDiff0 distance is greater than before clear and update    
                if agentPreyNeighDiff0>maxagentPreyNeighDiff:
                    maybeNeigh.clear()
                    maxagentPreyNeighDiff = agentPreyNeighDiff0
                    maybeNeigh.append(neighbour)
                # Node to keep the distance between Predator and Agent same
                if agentPreyNeighDiff0==maxagentPreyNeighDiff:
                    maybeNeigh.append(neighbour)
            simulatedAgentPos = random.choice(maybeNeigh)
            simulatedPreyNodeProb = updateSurveyPreyProd(size, simulatedAgentPos, simulatedPreyNodeProb, simulatedPreyPos)
            simulatedPreyPos = preyMovement(nodes, simulatedPreyPos)
            simulatedPreyNodeProb = updateTransitPreyProb(nodes, size, simulatedPreyNodeProb)
            timeStamp += 1
        nextNeigh.append(simulatedAgentPos)
    finalAgentPos = random.choice(nextNeigh)
    agentPath = random.choice(BFS(nodes, agentPos, finalAgentPos)["path"])
    
    if len(agentPath) > 1:
        agentPos = agentPath[1]
    return agentPos
    
def agent4Movement(nodes, size, predatorPos, agentPos, preyPos, preyNodeProb, preyCaught, probUse, distUse):
    """_summary_
        Function to make the Agent4 move based on the probabilistic decision graph received from the updateSurveyPreyProd function
    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        predatorPos (int): Current location of the predator on the graph
        agentPos (int): Current location of the Agent on the graph
        preyPos (int): Current location of the prey on the graph
        preyNodeProb (list): The list of the initialized probabilities for the entire graph

    Returns: The new location for the Agent to move to so that it can reach the goal node
        _type_: int
    """
    maybePrey = np.max(preyNodeProb)
    maybeprey = [i for i,prob in enumerate(preyNodeProb) if prob == maybePrey]
    maybePrey = random.choice(maybeprey)
    if maybePrey == preyPos:
        preyCaught += 1
    #SURVEY BELOW
    preyNodeProb = updateSurveyPreyProd(size, maybePrey, preyNodeProb, preyPos)
    maybePrey = np.max(preyNodeProb)
    maybeprey = [i for i,prob in enumerate(preyNodeProb) if prob == maybePrey]
    maybePrey = random.choice(maybeprey)
    maxPredDistNeigh = [agentPos]
    agentPathPred = random.choice(BFS(nodes, agentPos, predatorPos)["path"])
    agentPathPrey = random.choice(BFS(nodes, agentPos, maybePrey)["path"])
    agentPredDist = len(agentPathPred)
    agentPreyDist = len(agentPathPrey)
    maxPredDist = agentPredDist
    if agentPreyDist < agentPredDist:
        probUse += 1
        agentPos = simulateFuture(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist)
    else:
        distUse += 1
        for neighbour in nodes[agentPos]["neighbours"]:
            #print(neighbour, " maxPredDistNeigh", maxPredDistNeigh)
            # Find the distance between the Neighbour and the Prey
            predatorDict = BFS(nodes, neighbour, predatorPos)
            # Find the distance between the Neighbour and the Predator
            randInd = random.randint(0,len(predatorDict["path"])-1)
            newPathPred = predatorDict["path"][randInd]
            
            if len(newPathPred) > maxPredDist:
                maxPredDistNeigh.clear()
                maxPredDist = len(newPathPred)
                maxPredDistNeigh.append(neighbour)
                
            elif len(newPathPred) == maxPredDist:
                maxPredDistNeigh.append(neighbour)
        # Agent moves away from predator near to the prey
        agentPos = random.choice(maxPredDistNeigh)
    
    return agentPos, preyCaught, probUse, distUse


def agent4(nodes, size, predatorPos, agentPos, preyPos):
    """_summary_
        Function to actually make the agent move based on the new node coordinates 
        provided by agent3Movement function
    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        predatorPos (int): Current location of the predator on the graph
        agentPos (int): Current location of the Agent on the graph
        preyPos (int): Current location of the prey on the graph

    Returns: The result of the Agent movement which includes the status of completion, the step counter, 
        and the final agent path to completion
        _type_: json
    """
    threshold = 1000
    agentPath = list()
    agentPath.append(agentPos)
    predPath = list()
    predPath.append(predatorPos)
    preyCaught = 0
    probUse = 0
    distUse = 0
    preyNodeProb = generatePreyProb(size, agentPos) 
    for counter in range(1,threshold+1):
        # print("Counter: ", counter)
        # print("Agent: ", agentPos)
        # print("Prey: ", preyPos)
        # print("Predator: ", predatorPos)
        # print("========================")
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "preyCaught":preyCaught, "probUse":probUse, "distUse":distUse}
        
        if agentPos == predatorPos:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "preyCaught":preyCaught, "probUse":probUse, "distUse":distUse}
        
        agentPos, preyCaught, probUse, distUse = agent4Movement(nodes, size, predatorPos, agentPos, preyPos, preyNodeProb, preyCaught, probUse, distUse)
        agentPath.append(agentPos)
        
        preyNodeProb = updateSurveyPreyProd(size, agentPos, preyNodeProb, preyPos)
        # If Agent reaches Prey Position which ih the Goal State
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "preyCaught":preyCaught, "probUse":probUse, "distUse":distUse}
        # Making the prey move
        preyPos = preyMovement(nodes, preyPos)
        preyNodeProb = updateTransitPreyProb(nodes, size, preyNodeProb)
        # After the Prey movement takes place
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "preyCaught":preyCaught, "probUse":probUse, "distUse":distUse}
        # Conditions for Predator killing the agent
        predDict = predatorMovement(agentPos, predatorPos, nodes)
        if predDict["statusCode"] == 200:
            predatorPos = predDict["predatorPos"]
            predPath.append(predatorPos)
            
        elif predDict["statusCode"] == 400:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "preyCaught":preyCaught, "probUse":probUse, "distUse":distUse}
        
    return {"statusCode": 404, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "preyCaught":preyCaught, "probUse":probUse, "distUse":distUse}
        

def driver():
    """_summary_
        Driver Code for the Agent 4
    """
    nodes, size = genEnvironment()
    predatorPos, agentPos, preyPos = spawnCreatures()
    return agent4(nodes, size, predatorPos, agentPos, preyPos)

def dataCollection():
    """_summary_
        Function to collect the data regarding Agent 4, its performance and all other statistical information
    """
    final_data = list()
    for i in range(700):
        print("Counter: ",i)
        data = driver()
        final_data.append(data)
            
    df1 = pd.DataFrame(final_data)
    book = load_workbook('Agent4.xlsx')
    writer = pd.ExcelWriter('Agent4.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header = False)

    writer.save()
        
dataCollection()

#TEST FUNCTION TO TEST OUT BITS OF CODE AND FINALLY DRY RUN AGENT3
def testDriver():
    # nodes, size = genEnvironment()
    # predatorPos, agentPos, preyPos = spawnCreatures()
    # agent3data = agent4(nodes, size, predatorPos, agentPos, preyPos)     
    # print(agent3data)
    for i in range(10):
        data = driver()
        print(data)

#testDriver()

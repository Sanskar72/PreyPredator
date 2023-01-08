# Imports
from genenvironment import genEnvironment, spawnCreatures, preyMovement, BFS
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import random
import copy

def getMaxProb(NodeProb):
    maybePred = np.max(NodeProb)
    maybepred = [i for i,prob in enumerate(NodeProb) if prob == maybePred]
    return random.choice(maybepred)

def predatorMovement(agentPos, predatorPos, nodes):
    """_summary_
        Function for the movement of the Predator based on the Agent position. The Predator moves randomly to anyone of its neighbours with a probability of 0.4 and moves using the shortest path towards the Agent with a probability of 0.6
    Args:
        agentPos (int): The location of the agent on the graph
        predatorPos (int): The location of the predator on the graph
        nodes (Dictionary): Dictionary with all the node information in the graph

    Returns: The next move for the predator as well as the result if it becomes a success or Failure
        _type_: json
    """
    rand = random.random()
    if rand<0.6:
        bestNeigh = list()
        minLen = 50
        if agentPos != predatorPos:
            for neigh in nodes[predatorPos]["neighbours"]:
                # Find all the Shortest Paths 
                paths = BFS(nodes, neigh, agentPos)["path"]
                path = random.choice(paths)
                if len(bestNeigh)==0:
                    bestNeigh.append([neigh, len(path)])
                    minLen = len(path)
                elif len(path) == minLen:
                    bestNeigh.append([neigh, len(path)])
                elif len(path) < minLen:
                    bestNeigh.clear()
                    bestNeigh.append([neigh, len(path)])
                    minLen = len(path)
                #print("BN: ",bestNeigh)
            
            predArr = random.choice(bestNeigh)
            predatorPos = predArr[0]
            # Success
            return {"statusCode":200, "predatorPos":predatorPos}
            
        else:
            # Failure
            return {"statusCode": 400, "predatorPos":agentPos}

    else:
        # Make the predator move randomly
        predatorPos = random.choice(nodes[predatorPos]["neighbours"])
        return {"statusCode":200, "predatorPos":predatorPos}

def generatePredProb(size, predPos):
    """_summary_
        Function to initialise the probability of the Predator position on the graph.
        As we know the initial location of the Predator at the start we set that node probability as 1
        for all the other nodes it is 0
    Args:
        size (int): Length of the graph 
        agentPos (_type_): Location of the agent on the graph

    Returns: The probability list will all the node probabilities for it to be a Predator
        _type_: list
    """
    predNodeProb = [0]*size
    predNodeProb[predPos] = 1
    return predNodeProb

def updateTransitPredProb(nodes, size, predNodeProb, agentPos):
    """_summary_
        This Function is triggered after the predator moves which is done to update the belief states. Now, every node and its neighbour get updated based on the new information. 
        The Transition probability of the node is updated taking into consideration 
        the previous initialized Probabilities and the neighbours of the node into consideration
        
    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        predNodeProb (list): The list of the initialized probabilities for the entire graph for the predator
    
    Returns: The updated probability matrix for the entire graph maximizing the probability of the location of the prey
        _type_: list
    """
    newPredNodeProb = [0]*size
    transitGraph = dict()
    for node in range(size):
        bestNeigh = list()
        for neigh in nodes[node]["neighbours"]:
            paths = BFS(nodes, neigh, agentPos)["path"]
            path = random.choice(paths)
            if len(bestNeigh)==0:
                bestNeigh.append([neigh, len(path)])
                minLen = len(path)
            elif len(path) == minLen:
                bestNeigh.append([neigh, len(path)])
            elif len(path) < minLen:
                bestNeigh.clear()
                bestNeigh.append([neigh, len(path)])
                minLen = len(path)
                
        nextStep = [i[0] for i in bestNeigh]
        for step in nextStep:
            if transitGraph.get(step, False):
                if [node, 1/len(nextStep)] not in transitGraph[step]:
                    # Probability Formula used based on Conditional Probability
                    transitGraph[step].append([node, 1/len(nextStep)])
            else:
                transitGraph[step] = [[node, 1/len(nextStep)]]
    
    for node in range(size):
        if node in transitGraph.keys():
            for neigh in transitGraph[node]:
                newPredNodeProb[node] += 0.6*predNodeProb[neigh[0]]*neigh[1]
        else:
            newPredNodeProb[node] = 0
    #print("before other: ", np.sum(newPredNodeProb))
    for node in range(size):
        sum = 0
        for neigh in nodes[node]["neighbours"]:
            sum += 0.4*(1/nodes[neigh]["degree"])*predNodeProb[neigh] 
        newPredNodeProb[node] += sum
        #print(newPredNodeProb)


    #print("transitGraph: ",transitGraph)
    return newPredNodeProb

def updateSurveyPredProd(size, surveySpot, predNodeProb, predPos):
    """_summary_
        After surveying the graph for the predator we keep updating the probabilities of each node based on the conditional probability

    Args:
        size (_type_): Length of the graph
        surveySpot (int): Location with the highest probability of having the prey used for further calculation of Agent movement
        preyNodeProb (list): The list of the initialized probabilities for the entire graph
        preyPos (int): Current location of the prey on the graph

    Returns: 
        Surveying all the nodes in the graph we geth the final probability matrix using conditional Probability to predict the prey location
    """
    # Success Condition for survey
    if surveySpot == predPos:
        newPredNodeProb = [0]*size
        newPredNodeProb[surveySpot] = 1
        return newPredNodeProb
    # If the survey probability is not 1 we update the belief states oof every node we survey 
    elif predNodeProb[surveySpot] != 1:
        newPredNodeProb = [-100]*size
        for i in range(size):
            if i == surveySpot:
                newPredNodeProb[i] = 0
                continue
            newPredNodeProb[i] = predNodeProb[i]/(1-predNodeProb[surveySpot])
        return newPredNodeProb
    else:
        newPredNodeProb = [0]*size
        newPredNodeProb[surveySpot] = 1
        return newPredNodeProb

def simulateFuture(nodes, agentPos, preyPos, agentPreyDist):
    """_summary_
        The Function is used to simulate the entire process after surveying for belief states of each nodes. we simulate 
    Args:
        nodes (Dictionary): Dictionary with all the node information in the graph
        predatorPos (int): The location of the Predator on the chain graph
        agentPos (int): The location of the Agent on the chain graph
        preyPos (int): The location of the Prey on the chain graph

    Returns: It returns the Agent position after simulating the entire process 10 times and then picking one at random
        _type_: int
    """
    nextNeigh = list()
    simulationCount = 10
    for _ in range(simulationCount):
        simulatedAgentPos = copy.deepcopy(agentPos)
        simulatedPreyPos = copy.deepcopy(preyPos)
        timeStamp = 0
        while timeStamp<agentPreyDist and simulatedAgentPos != simulatedPreyPos:
            maybeNeigh = [simulatedAgentPos]
            # Find the shortest path between Prey and Agent
            agentPathPrey = random.choice(BFS(nodes, simulatedAgentPos, simulatedPreyPos)["path"])
            agentPreyDist = len(agentPathPrey)
            maxagentPreyNeighDiff = 0
            for neighbour in nodes[simulatedAgentPos]["neighbours"]:
                preyDict = BFS(nodes, neighbour, simulatedPreyPos)
                randInd = random.randint(0,len(preyDict["path"])-1)
                newPathPrey = preyDict["path"][randInd]
                agentPreyNeighDiff0 = (agentPreyDist-len(newPathPrey))
                # If agentPreyNeighDiff0 distance is greater than before clear and update    
                if agentPreyNeighDiff0>maxagentPreyNeighDiff:
                    maybeNeigh.clear()
                    maxagentPreyNeighDiff = agentPreyNeighDiff0
                    maybeNeigh.append(neighbour)
                # Node to keep the distance between Predator and Agent same    
                if agentPreyNeighDiff0==maxagentPreyNeighDiff:
                    maybeNeigh.append(neighbour)
            simulatedAgentPos = random.choice(maybeNeigh)
            simulatedPreyPos = preyMovement(nodes, simulatedPreyPos)
            timeStamp += 1
        nextNeigh.append(simulatedAgentPos)
    finalAgentPos = random.choice(nextNeigh)
    agentPath = random.choice(BFS(nodes, agentPos, finalAgentPos)["path"])
    
    if len(agentPath) > 1:
        agentPos = agentPath[1]
    else:
        agentPos = random.choice(BFS(nodes, agentPos, preyPos)["path"])[1]
    return agentPos

def agent6Movement(nodes, size, predPos, agentPos, preyPos, predNodeProb, predCaught, probUse, distUse):
    """_summary_
        Function to make the Agent3 based on the probabilistic decision graph received from the updateSurveyPredProd function

    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        predatorPos (int): Current location of the predator on the graph
        agentPos (int): Current location of the Agent on the graph
        preyPos (int): Current location of the prey on the graph
        preyNodeProb (list): The list of the initialized probabilities for the entire graph

    Returns: The new location for the Agent to move to so that it can reach the goal node
        _type_: int
    """
    maybePred = getMaxProb(predNodeProb)
    #SURVEY BELOW
    predNodeProb = updateSurveyPredProd(size, maybePred, predNodeProb, predPos)
    maybePred = getMaxProb(predNodeProb)
    if maybePred == predPos:
        predCaught += 1
    maxPredDistNeigh = [agentPos]
    agentPathPred = random.choice(BFS(nodes, agentPos, maybePred)["path"])
    agentPathPrey = random.choice(BFS(nodes, agentPos, preyPos)["path"])
    agentPredDist = len(agentPathPred)
    agentPreyDist = len(agentPathPrey)
    maxPredDist = agentPredDist
    if agentPreyDist < agentPredDist*0.7:
        probUse += 1
        agentPos = simulateFuture(nodes, agentPos, preyPos, agentPreyDist)
    else:
        distUse += 1
        for neighbour in nodes[agentPos]["neighbours"]:
            #print(neighbour, " maxPredDistNeigh", maxPredDistNeigh)
            predatorDict = BFS(nodes, neighbour, maybePred)
            # Find the distance between the Neighbour and the Predator
            randInd = random.randint(0,len(predatorDict["path"])-1)
            newPathPred = predatorDict["path"][randInd]
            
            if len(newPathPred) > maxPredDist:
                maxPredDistNeigh.clear()
                maxPredDist = len(newPathPred)
                maxPredDistNeigh.append(neighbour)
                
            elif len(newPathPred) == maxPredDist:
                maxPredDistNeigh.append(neighbour)
        # Agent moves away from predator near to the prey
        agentPos = random.choice(maxPredDistNeigh)
    
    return agentPos, predCaught, probUse, distUse

def agent6(nodes, size, predatorPos, agentPos, preyPos):
    """_summary_
        Function to actually make the agent move based on the new node coordinates 
        provided by agent5Movement function
    Args:
        nodes (Dictionary with all the node information in the graph
        size (int): Length of the graph 
        predatorPos (int): Current location of the predator on the graph
        agentPos (int): Current location of the Agent on the graph
        preyPos (int): Current location of the prey on the graph

    Returns: The result of the Agent movement which includes the status of completion, the step counter, 
        and the final agent path to completion
        _type_: json
    """
    threshold = 1000
    agentPath = list()
    agentPath.append(agentPos)
    predPath = list()
    predPath.append(predatorPos)
    preyPath = list()
    preyPath.append(preyPos)
    predNodeProb = generatePredProb(size, predatorPos) 
    predCaught = 0
    probUse = 0
    distUse = 0
    for counter in range(1,threshold+1):
        # print("Counter: ", counter)
        # print("Agent: ", agentPos)
        # print("Prey: ", preyPos)
        # print("Predator: ", predatorPos)
        # print("========================")
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        
        if agentPos == predatorPos:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        
        agentPos, predCaught, probUse, distUse = agent6Movement(nodes, size, predatorPos, agentPos, preyPos, predNodeProb, predCaught, probUse, distUse)
        agentPath.append(agentPos)
        predNodeProb = updateSurveyPredProd(size, agentPos, predNodeProb, predatorPos)
        #print("Sum of predNodeProb: ", np.sum(predNodeProb))
        # If Agent reaches Prey Position which ih the Goal State
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        # Making the prey move
        preyPos = preyMovement(nodes, preyPos)
        preyPath.append(preyPos)
        # After the Prey movement takes place
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        # Conditions for Predator killing the agent
        predDict = predatorMovement(agentPos, predatorPos, nodes)
        if predDict["statusCode"] == 200:
            predatorPos = predDict["predatorPos"]
            predPath.append(predatorPos)
            
        elif predDict["statusCode"] == 400:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        
        predNodeProb = updateTransitPredProb(nodes, size, predNodeProb, agentPos)
        #print("Sum of predNodeProb: ", np.sum(predNodeProb))
        
    return {"statusCode": 404, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}


def driver():
    """_summary_
        Driver Code for the Agent 6
    """
    nodes, size = genEnvironment()
    predatorPos, agentPos, preyPos = spawnCreatures()
    return agent6(nodes, size, predatorPos, agentPos, preyPos)

def dataCollection():
    """_summary_
        Function to collect the data regarding Agent 6, its performance and all other statistical information
    """
    final_data = list()
    for i in range(300):
        print("Counter: ",i)
        data = driver()
        final_data.append(data)
            
    df1 = pd.DataFrame(final_data)
    book = load_workbook('Agent6.xlsx')
    writer = pd.ExcelWriter('Agent6.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header = False)

    writer.save()

dataCollection()

#TEST FUNCTION TO TEST OUT BITS OF CODE AND FINALLY DRY RUN AGENT3
def testDriver():
    nodes, size = genEnvironment()
    predatorPos, agentPos, preyPos = spawnCreatures()
    predNodeProb = generatePredProb(size, agentPos) 
    for i in range(10):
        predNodeProb = updateTransitPredProb(nodes, size, predNodeProb, agentPos)
        print("Sum of predNodeProb: ", np.sum(predNodeProb))
        print("================================")
    #agent5data = agent5(nodes, size, predatorPos, agentPos, preyPos)     
    #print(agent5data)

#testDriver()
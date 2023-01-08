from genenvironment import genEnvironment, spawnCreatures, preyMovement, BFS
import pandas as pd
from openpyxl import load_workbook
import numpy as np
import random
import copy

def simulateFuturePrey(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist):
    """_summary_
        The Function is used to simulate the entire process after surveying for belief states of each nodes. we simulate 
    Args:
        nodes (Dictionary): Dictionary with all the node information in the graph
        predatorPos (int): The location of the Predator on the chain graph
        agentPos (int): The location of the Agent on the chain graph
        preyPos (int): The location of the Prey on the chain graph

    Returns: It returns the Agent position after simulating the entire process 10 times and then picking one at random
        _type_: int
    """
    nextPrey = list()
    simulationCount = 30
    for _ in range(simulationCount):
        simulatedAgentPos = copy.deepcopy(agentPos)
        simulatedPreyNodeProb = copy.deepcopy(preyNodeProb)
        simulatedPreyPos = copy.deepcopy(preyPos)
        timeStamp = 0
        while timeStamp<agentPreyDist and simulatedAgentPos != simulatedPreyPos:
            maybeNeigh = [simulatedAgentPos]
            maybePrey = np.max(simulatedPreyNodeProb)
            maybeprey = [i for i,prob in enumerate(simulatedPreyNodeProb) if prob == maybePrey]
            maybePrey = random.choice(maybeprey)
            # Survey and create the belief States for the Future Simulations
            simulatedPreyNodeProb = updateSurveyPreyProd(size, maybePrey, simulatedPreyNodeProb, simulatedPreyPos)
            maybePrey = np.max(simulatedPreyNodeProb)
            maybeprey = [i for i,prob in enumerate(simulatedPreyNodeProb) if prob == maybePrey]
            maybePrey = random.choice(maybeprey)
            agentPathPrey = random.choice(BFS(nodes, simulatedAgentPos, maybePrey)["path"])
            agentPreyDist = len(agentPathPrey)
            maxagentPreyNeighDiff = 0
            for neighbour in nodes[simulatedAgentPos]["neighbours"]:
                # Find the path between the Neighbour and the Prey
                preyDict = BFS(nodes, neighbour, maybePrey)
                randInd = random.randint(0,len(preyDict["path"])-1)
                newPathPrey = preyDict["path"][randInd]
                agentPreyNeighDiff0 = (agentPreyDist-len(newPathPrey))
                if agentPreyNeighDiff0>maxagentPreyNeighDiff:
                    maybeNeigh.clear()
                    maxagentPreyNeighDiff = agentPreyNeighDiff0
                    maybeNeigh.append(neighbour)
                    
                if agentPreyNeighDiff0==maxagentPreyNeighDiff:
                    maybeNeigh.append(neighbour)
            # Survey and create the belief States for the Future Simulations
            simulatedAgentPos = random.choice(maybeNeigh)
            simulatedPreyNodeProb = updateSurveyPreyProd(size, simulatedAgentPos, simulatedPreyNodeProb, simulatedPreyPos)
            simulatedPreyPos = preyMovement(nodes, simulatedPreyPos)
            simulatedPreyNodeProb = updateTransitPreyProb(nodes, size, simulatedPreyNodeProb)
            timeStamp += 1
        maybePrey = np.max(simulatedPreyNodeProb)
        maybeprey = [i for i,prob in enumerate(simulatedPreyNodeProb) if prob == maybePrey]
        maybePrey = random.choice(maybeprey)
        nextPrey.append(maybePrey)
    probPreyPos = random.choice(nextPrey)
    
    return probPreyPos
        

def predatorMovement(agentPos, predatorPos, nodes):
    """_summary_
        Function for the movement of the Predator based on the Agent position
    Args:
        agentPos (int): The location of the agent on the graph
        predatorPos (int): The location of the predator on the graph
        nodes (Dictionary): Dictionary with all the node information in the graph

    Returns: The next move for the predator as well as the result if it becomes a success or Failure
        _type_: json
    """
    rand = random.random()
    if rand<0.6:
        bestNeigh = list()
        minLen = 50
        if agentPos != predatorPos:
            for neigh in nodes[predatorPos]["neighbours"]:
                # Find all the Shortest Paths 
                paths = BFS(nodes, neigh, agentPos)["path"]
                path = random.choice(paths)
                if len(bestNeigh)==0:
                    bestNeigh.append([neigh, len(path)])
                    minLen = len(path)
                elif len(path) == minLen:
                    bestNeigh.append([neigh, len(path)])
                elif len(path) < minLen:
                    bestNeigh.clear()
                    bestNeigh.append([neigh, len(path)])
                    minLen = len(path)
                #print("BN: ",bestNeigh)
            
            predArr = random.choice(bestNeigh)
            predatorPos = predArr[0]
            return {"statusCode":200, "predatorPos":predatorPos}
            
        else:
            return {"statusCode": 400, "predatorPos":agentPos}

    else:
        predatorPos = random.choice(nodes[predatorPos]["neighbours"])
        return {"statusCode":200, "predatorPos":predatorPos}

def generatePreyProb(size, agentPos):
    """_summary_
        Function to initialize the Prey node position probability as 0 and the initial probability of 1/(size of graph -1) to every node 
        in the graph 

    Args:
        size (int): Length of the graph 
        agentPos (int): Location of the agent on the graph

    Returns: A list of the initialized probabilities for the entire graph
        _type_: list
    """
    preyNodeProb = list()
    for i in range(size):
        if i == agentPos:
            preyNodeProb.append(0)
            continue
        preyNodeProb.append(1/(size-1))
    return preyNodeProb

def updateTransitPreyProb(nodes, size, preyNodeProb):
    """_summary_
        This Function is triggered after the prey moves which is done to update the belief states. Now, every node and its neighbour get updated based on the new information. 
        The Transition probability of the node is updated taking into consideration 
        the previous initialized Probabilities and the neighbours of the node into consideration
        
    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        predNodeProb (list): The list of the initialized probabilities for the entire graph for the predator
    
    Returns: The updated probability matrix for the entire graph maximizing the probability of the location of the prey
        _type_: list
    """
    newPreyNodeProb = [-100]*size
    for node in range(size):
        prob = preyNodeProb[node]/(nodes[node]["degree"]+1)
        for i in range(nodes[node]["degree"]):
            neighbour = nodes[node]["neighbours"][i]
            prob += preyNodeProb[neighbour]/(nodes[neighbour]["degree"]+1)
            newPreyNodeProb[node] = prob
            
    return newPreyNodeProb

def updateSurveyPreyProd(size, surveySpot, preyNodeProb, preyPos):
    """_summary_
        After surveying the graph for the prey we keep updating the probabilities of each node based on the conditional probability
    Args:
        size (int): Length of the graph
        surveySpot (int): Location with the highest probability of having the prey used for further calculation of Agent movement
        preyNodeProb (list): The list of the initialized probabilities for the entire graph
        preyPos (int): Current location of the prey on the graph

    Returns: 
        Surveying all the nodes in the graph we geth the final probability matrix using conditional Probability to predict the prey location
    """
    defect = np.random.rand()
    if surveySpot == preyPos and defect>0.1:
        for i in range(size):
            if i == surveySpot:
                preyNodeProb[i] = 1
                continue
            preyNodeProb[i] = 0
        return preyNodeProb
    else:
        newPreyNodeProb = [-100]*size
        for i in range(size):
            if i == surveySpot:
                newPreyNodeProb[i] = 0.1*preyNodeProb[surveySpot]/(1-preyNodeProb[surveySpot]+(0.1*preyNodeProb[surveySpot]))
                continue
            newPreyNodeProb[i] = preyNodeProb[i]/(1-preyNodeProb[surveySpot]+(0.1*preyNodeProb[surveySpot]))
        return newPreyNodeProb

def generatePredProb(size, predPos):
    """_summary_
        Function to initialise the probability of the Predator position on the graph.
        As we know the initial location of the Predator at the start we set that node probability as 1
        for all the other nodes it is 0
    Args:
        size (int): Length of the graph 
        agentPos (int): Location of the agent on the graph

    Returns: The probability list will all the node probabilities for it to be a Predator
        _type_: list
    """
    predNodeProb = [0]*size
    predNodeProb[predPos] = 1
    return predNodeProb

def updateTransitPredProb(nodes, size, predNodeProb, agentPos):
    """_summary_       
        Function to survey each and every node to find out the probability of that node being the Predator position

    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        preyNodeProb (list): The list of the initialized probabilities for the entire graph
    
    Returns:The updated probability matrix for the entire graph maximizing the probability of the location of the predator
        _type_: list
    """
    newPredNodeProb = [0]*size
    transitGraph = dict()
    for node in range(size):
        bestNeigh = list()
        for neigh in nodes[node]["neighbours"]:
            paths = BFS(nodes, neigh, agentPos)["path"]
            path = random.choice(paths)
            if len(bestNeigh)==0:
                bestNeigh.append([neigh, len(path)])
                minLen = len(path)
            elif len(path) == minLen:
                bestNeigh.append([neigh, len(path)])
            elif len(path) < minLen:
                bestNeigh.clear()
                bestNeigh.append([neigh, len(path)])
                minLen = len(path)
                
        nextStep = [i[0] for i in bestNeigh]
        for step in nextStep:
            if transitGraph.get(step, False):
                if [node, 1/len(nextStep)] not in transitGraph[step]:
                    # Probability Formula used based on Conditional Probability
                    transitGraph[step].append([node, 1/len(nextStep)])
            else:
                transitGraph[step] = [[node, 1/len(nextStep)]]
    
    for node in range(size):
        if node in transitGraph.keys():
            for neigh in transitGraph[node]:
                newPredNodeProb[node] += 0.6*predNodeProb[neigh[0]]*neigh[1]
        else:
            newPredNodeProb[node] = 0
    # print("before other: ", np.sum(newPredNodeProb))
    for node in range(size):
        sum = 0
        for neigh in nodes[node]["neighbours"]:
            sum += 0.4*(1/nodes[neigh]["degree"])*predNodeProb[neigh] 
        newPredNodeProb[node] += sum
        #print(newPredNodeProb)
    #print("transitGraph: ",transitGraph)
    return newPredNodeProb

def updateSurveyPredProd(size, surveySpot, predNodeProb, predPos):
    """_summary_
        After surveying the graph for the prey we keep updating the probabilities of each node based on the conditional probability

    Args:
        size (int): Length of the graph
        surveySpot (int): Location with the highest probability of having the prey used for further calculation of Agent movement
        preyNodeProb (list): The list of the initialized probabilities for the entire graph
        preyPos (int): Current location of the prey on the graph

    Returns: 
        Surveying all the nodes in the graph we geth the final probability matrix using conditional Probability to predict the prey location
    """
    defect = np.random.rand()
    if surveySpot == predPos and defect>0.1:
        newPredNodeProb = [0]*size
        newPredNodeProb[surveySpot] = 1
        return newPredNodeProb
    elif predNodeProb[surveySpot] != 1:
        newPredNodeProb = [-100]*size
        for i in range(size):
            if i == surveySpot:
                newPredNodeProb[i] = 0.1*predNodeProb[surveySpot]/(1-predNodeProb[surveySpot]+(0.1*predNodeProb[surveySpot]))
                continue
            newPredNodeProb[i] = predNodeProb[i]/(1-predNodeProb[surveySpot]+(0.1*predNodeProb[surveySpot]))
        return newPredNodeProb
    else:
        newPredNodeProb = [0]*size
        newPredNodeProb[surveySpot] = 1
        return newPredNodeProb

def simulateFuture(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist):
    """_summary_
        The Function is used to simulate the entire process after surveying for belief states of each nodes. we simulate 
    Args:
        nodes (Dictionary): Dictionary with all the node information in the graph
        predatorPos (int): The location of the Predator on the chain graph
        agentPos (int): The location of the Agent on the chain graph
        preyPos (int): The location of the Prey on the chain graph

    Returns: It returns the Agent position after simulating the entire process 10 times and then picking one at random
        _type_: int
    """
    nextNeigh = list()
    simulationCount = 30
    for _ in range(simulationCount):
        simulatedAgentPos = copy.deepcopy(agentPos)
        simulatedPreyNodeProb = copy.deepcopy(preyNodeProb)
        simulatedPreyPos = copy.deepcopy(preyPos)
        timeStamp = 0
        while timeStamp<agentPreyDist and simulatedAgentPos != simulatedPreyPos:
            maybeNeigh = [simulatedAgentPos]
            maybePrey = np.max(simulatedPreyNodeProb)
            maybeprey = [i for i,prob in enumerate(simulatedPreyNodeProb) if prob == maybePrey]
            maybePrey = random.choice(maybeprey)
            simulatedPreyNodeProb = updateSurveyPreyProd(size, maybePrey, simulatedPreyNodeProb, simulatedPreyPos)
            maybePrey = np.max(simulatedPreyNodeProb)
            maybeprey = [i for i,prob in enumerate(simulatedPreyNodeProb) if prob == maybePrey]
            maybePrey = random.choice(maybeprey)
            agentPathPrey = random.choice(BFS(nodes, simulatedAgentPos, maybePrey)["path"])
            agentPreyDist = len(agentPathPrey)
            maxagentPreyNeighDiff = 0
            for neighbour in nodes[simulatedAgentPos]["neighbours"]:
                preyDict = BFS(nodes, neighbour, maybePrey)
                randInd = random.randint(0,len(preyDict["path"])-1)
                newPathPrey = preyDict["path"][randInd]
                agentPreyNeighDiff0 = (agentPreyDist-len(newPathPrey))
                if agentPreyNeighDiff0>maxagentPreyNeighDiff:
                    maybeNeigh.clear()
                    maxagentPreyNeighDiff = agentPreyNeighDiff0
                    maybeNeigh.append(neighbour)
                    
                if agentPreyNeighDiff0==maxagentPreyNeighDiff:
                    maybeNeigh.append(neighbour)
            simulatedAgentPos = random.choice(maybeNeigh)
            simulatedPreyNodeProb = updateSurveyPreyProd(size, simulatedAgentPos, simulatedPreyNodeProb, simulatedPreyPos)
            simulatedPreyPos = preyMovement(nodes, simulatedPreyPos)
            simulatedPreyNodeProb = updateTransitPreyProb(nodes, size, simulatedPreyNodeProb)
            timeStamp += 1
        nextNeigh.append(simulatedAgentPos)
    finalAgentPos = random.choice(nextNeigh)
    agentPath = random.choice(BFS(nodes, agentPos, finalAgentPos)["path"])
    
    if len(agentPath) > 1:
        agentPos = agentPath[1]
    return agentPos
        
def agent8Movement(nodes, size, predPos, agentPos, preyPos, predNodeProb, preyNodeProb, preyCaught, predCaught, probUse, distUse):
    """_summary_
        Function to make the Agent 8 bonus move based on the probabilistic decision
    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        predatorPos (int): Current location of the predator on the graph
        agentPos (int): Current location of the Agent on the graph
        preyPos (int): Current location of the prey on the graph
        preyNodeProb (list): The list of the initialized probabilities for the entire graph
    Returns: The new location for the Agent to move to so that it can reach the goal node
        _type_: int
    """

    
    maybePred = np.max(predNodeProb)
    maybePrey = np.max(preyNodeProb)
    maybeprey = [i for i,prob in enumerate(preyNodeProb) if prob == maybePrey]
    maybePrey = random.choice(maybeprey)
    agentPathPrey = random.choice(BFS(nodes, agentPos, maybePrey)["path"])
    agentPreyDist = len(agentPathPrey)
    maybePrey = simulateFuturePrey(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist)
    if maybePred < 0.5:
        maybepred = [i for i,prob in enumerate(predNodeProb) if prob == maybePred]
        maybePred = random.choice(maybepred)
        if maybePred == predPos:
            predCaught += 1
        if maybePrey == preyPos:
            preyCaught += 1
        #SURVEY BELOW
        predNodeProb = updateSurveyPredProd(size, maybePred, predNodeProb, predPos)
        preyNodeProb = updateSurveyPreyProd(size, maybePred, preyNodeProb, preyPos)
        maybePred = np.max(predNodeProb)
        maybepred = [i for i,prob in enumerate(predNodeProb) if prob == maybePred]
        maybePred = random.choice(maybepred)
        maybePrey = np.max(preyNodeProb)
        maybeprey = [i for i,prob in enumerate(preyNodeProb) if prob == maybePrey]
        maybePrey = random.choice(maybeprey)
        agentPathPrey = random.choice(BFS(nodes, agentPos, maybePrey)["path"])
        agentPreyDist = len(agentPathPrey)
        maybePrey = simulateFuturePrey(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist)

    else:
        maybePrey = np.max(preyNodeProb)
        maybeprey = [i for i,prob in enumerate(preyNodeProb) if prob == maybePrey]
        maybePrey = random.choice(maybeprey)
        agentPathPrey = random.choice(BFS(nodes, agentPos, maybePrey)["path"])
        agentPreyDist = len(agentPathPrey)
        maybePrey = simulateFuturePrey(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist)
        if maybePred == predPos:
            predCaught += 1
        if maybePrey == preyPos:
            preyCaught += 1
        #SURVEY BELOW
        preyNodeProb = updateSurveyPreyProd(size, maybePrey, preyNodeProb, preyPos)
        predNodeProb = updateSurveyPredProd(size, maybePrey, predNodeProb, predPos)
        maybePred = np.max(predNodeProb)
        maybepred = [i for i,prob in enumerate(predNodeProb) if prob == maybePred]
        maybePred = random.choice(maybepred)

        maybePrey = np.max(preyNodeProb)
        maybeprey = [i for i,prob in enumerate(preyNodeProb) if prob == maybePrey]
        maybePrey = random.choice(maybeprey)
        agentPathPrey = random.choice(BFS(nodes, agentPos, maybePrey)["path"])
        agentPreyDist = len(agentPathPrey)
        maybePrey = simulateFuturePrey(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist)

    # print(maybePrey, maybePred)
    # print(preyNodeProb[maybePrey], predNodeProb[maybePred])
    # print(preyPos, predPos)
    # print(preyNodeProb[preyPos], predNodeProb[predPos])
    # print("======================================================")
    maxPredDistNeigh = [agentPos]
    agentPathPred = random.choice(BFS(nodes, agentPos, maybePred)["path"])
    agentPathPrey = random.choice(BFS(nodes, agentPos, maybePrey)["path"])
    agentPredDist = len(agentPathPred)
    agentPreyDist = len(agentPathPrey)
    maxPredDist = agentPredDist
    if agentPreyDist < agentPredDist*0.7:
        probUse += 1
        agentPos = simulateFuture(nodes, size, agentPos, preyPos, preyNodeProb, agentPreyDist)
    else:
        distUse += 1
        for neighbour in nodes[agentPos]["neighbours"]:
            #print(neighbour, " maxPredDistNeigh", maxPredDistNeigh)
            predatorDict = BFS(nodes, neighbour, maybePred)
            # Find the distance between the Neighbour and the Predator
            randInd = random.randint(0,len(predatorDict["path"])-1)
            newPathPred = predatorDict["path"][randInd]
            
            if len(newPathPred) > maxPredDist:
                maxPredDistNeigh.clear()
                maxPredDist = len(newPathPred)
                maxPredDistNeigh.append(neighbour)
                
            elif len(newPathPred) == maxPredDist:
                maxPredDistNeigh.append(neighbour)
        # Agent moves away from predator near to the prey
        agentPos = random.choice(maxPredDistNeigh)
    
    return agentPos, preyCaught, predCaught, probUse, distUse
    
def agent8(nodes, size, predatorPos, agentPos, preyPos):
    """_summary_
        Function to actually make the agent move based on the new node coordinates 
        provided by agent3Movement function
    Args:
        nodes (Dictionary):  Dictionary with all the node information in the graph
        size (int): Length of the graph 
        predatorPos (int): Current location of the predator on the graph
        agentPos (int): Current location of the Agent on the graph
        preyPos (int): Current location of the prey on the graph

    Returns: The result of the Agent movement which includes the status of completion, the step counter, 
        and the final agent path to completion
        _type_: json
    """
    threshold = 1000
    agentPath = list()
    agentPath.append(agentPos)
    predPath = list()
    predPath.append(predatorPos)
    preyPath = list()
    preyPath.append(preyPos)
    preyCaught, predCaught = 0, 0
    probUse, distUse = 0, 0
    predNodeProb = generatePredProb(size, predatorPos) 
    preyNodeProb = generatePreyProb(size, agentPos) 
    for counter in range(1,threshold+1):
        # print("Counter: ", counter)
        # print("Agent: ", agentPos)
        # print("Prey: ", preyPos)
        # print("Predator: ", predatorPos)
        # print("========================")
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "preyCaught":preyCaught, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        
        if agentPos == predatorPos:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "preyCaught":preyCaught, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        
        agentPos, preyCaught, predCaught, probUse, distUse = agent8Movement(nodes, size, predatorPos, agentPos, preyPos, predNodeProb, preyNodeProb, preyCaught, predCaught, probUse, distUse)
        agentPath.append(agentPos)
        predNodeProb = updateSurveyPredProd(size, agentPos, predNodeProb, predatorPos)
        preyNodeProb = updateSurveyPreyProd(size, agentPos, preyNodeProb, preyPos)
        #print("Sum of predNodeProb: ", np.sum(predNodeProb))
        # If Agent reaches Prey Position which ih the Goal State
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "preyCaught":preyCaught, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        # Making the prey move
        preyPos = preyMovement(nodes, preyPos)
        preyPath.append(preyPos)
        preyNodeProb = updateTransitPreyProb(nodes, size, preyNodeProb)
        # After the Prey movement takes place
        if agentPos == preyPos:
            return {"statusCode": 200, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "preyCaught":preyCaught, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        # Conditions for Predator killing the agent
        predDict = predatorMovement(agentPos, predatorPos, nodes)
        if predDict["statusCode"] == 200:
            predatorPos = predDict["predatorPos"]
            predPath.append(predatorPos)
            
        elif predDict["statusCode"] == 400:
            return {"statusCode": 400, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "preyCaught":preyCaught, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        #print("Predpos: ", predatorPos)
        predNodeProb = updateTransitPredProb(nodes, size, predNodeProb, agentPos)
        #print("Max of predNodeProb: ", np.max(predNodeProb))
        #print("Sum of predNodeProb: ", np.sum(predNodeProb))
        #print("======================")
        
    return {"statusCode": 404, "steps":counter, "AgentPath":agentPath, "PredPath":predPath, "PreyPath":preyPath, "preyCaught":preyCaught, "predCaught":predCaught, "probUse":probUse, "distUse":distUse}
        
def driver():
    """_summary_
        Driver Code for the Agent 8 bonus
    """
    nodes, size = genEnvironment()
    predatorPos, agentPos, preyPos = spawnCreatures()
    return agent8(nodes, size, predatorPos, agentPos, preyPos)

def dataCollection():
    """_summary_
        Function to collect the data regarding Agent 8 bonus, its performance and all other statistical information
    """
    final_data = list()
    for i in range(300):
        print("Counter: ",i)
        data = driver()
        final_data.append(data)
            
    df1 = pd.DataFrame(final_data)
    book = load_workbook('Agent8_1.xlsx')
    writer = pd.ExcelWriter('Agent8_1.xlsx', engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}

    for sheetname in writer.sheets:
        df1.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False, header = False)

    writer.save()
    
dataCollection()

#TEST FUNCTION TO TEST OUT BITS OF CODE AND FINALLY DRY RUN AGENT3
def testDriver():
    nodes, size = genEnvironment()
    predPos, agentPos, preyPos = spawnCreatures()
    preyNodeProb = generatePreyProb(size, agentPos)
    predNodeProb = generatePredProb(size, predPos)
    for i in range(10):
        print(np.sum(preyNodeProb))
        print(np.sum(predNodeProb))
        print("================================")
        predNodeProb = updateSurveyPredProd(size, np.argmax(predNodeProb), predNodeProb, predPos)
        preyNodeProb = updateSurveyPreyProd(size, np.argmax(preyNodeProb), preyNodeProb, preyPos)
        preyPos = preyMovement(nodes, preyPos)
        predDict = predatorMovement(agentPos, predPos, nodes)
        if predDict["statusCode"] == 200:
            predPos = predDict["predatorPos"]
        preyNodeProb = updateTransitPreyProb(nodes, size, preyNodeProb)
        predNodeProb = updateTransitPredProb(nodes, size, predNodeProb, agentPos)
        
#testDriver()